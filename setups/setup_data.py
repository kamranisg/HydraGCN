import torch
import numpy as np
import os
from openpyxl import load_workbook

from base.datasets.dataset_setup import DatasetSetupBuilder

def create_dir(dir_):
    if not os.path.isdir(dir_):
        print('Create directory', dir_)
        os.mkdir(dir_)


class Covid19ICU(DatasetSetupBuilder):
    def __init__(self):
        super(Covid19ICU, self).__init__()

    def build_dataset(self):
        if not os.path.isdir('../data/COVID19_ICU'):
            print('Creating COVID19_ICU folder structure ...')
            file = '../sources/COVID19/COVID19_ICU.xlsx'
            data = load_workbook(file)
            data_work = data['COVID19']

            feature_list = ['B', 'W', 'AD', 'BU', 'BV', 'DH', 'DI', 'DJ', 'DK', 'DL', 'ES',
                            'ET', 'EU', 'EV', 'EW', 'EX', 'EY', 'EZ', 'FA', 'FB']

            meta_list = ['B', 'BU', 'BV']

            target = 'M'

            features = []
            meta_data = []
            targets = []
            feature_names = []

            for j, para in enumerate(feature_list):
                feature_names.append(data_work[para + str(1)].value)

            idx = 2
            while data_work['A' + str(idx)].value is not None:
                curr_feat = []
                curr_meta = []
                if data_work[target + str(idx)].value is not None:
                    targets.append(data_work[target + str(idx)].value)
                else:
                    idx += 1
                    continue
                for j, para in enumerate(feature_list):
                    if data_work[para + str(idx)].value is not None:
                        val = data_work[para + str(idx)].value
                        curr_feat.append(np.round(val, 3))
                    else:
                        curr_feat.append(-1)
                features.append(curr_feat)
                for j, para in enumerate(meta_list):
                    if data_work[para + str(idx)].value is not None and not data_work[para + str(idx)].value == -1:
                        curr_meta.append(data_work[para + str(idx)].value)
                    else:
                        curr_meta.append(-1)
                meta_data.append(curr_meta)
                idx += 1

            features = np.array(features)
            meta_data = np.array(meta_data)

            for i in range(features.shape[1]):
                mask = features[:, i] == -1
                features[mask, i] = features.mean(0)[i]

            for i in range(meta_data.shape[1]):
                mask = meta_data[:, i] == -1
                meta_data[mask, i] = meta_data.mean(0)[i]

            create_dir('../data')
            create_dir('../data/COVID19_ICU')
            np.save('../data/COVID19_ICU/features.npy', features)
            np.save('../data/COVID19_ICU/targets.npy', targets)
            np.save('../data/COVID19_ICU/meta_data.npy', meta_data)
            np.save('../data/COVID19_ICU/feature_names.npy', feature_names)

            print('COVID19_ICU folder structure created, proceeding ...')
        else:
            print('COVID19_ICU folder structure already created, proceeding ...')

        dirs = {'img': None,
                'feat': '../data/COVID19_ICU',
                'targets': ['../data/COVID19_ICU'],
                'loss': [torch.nn.BCEWithLogitsLoss],
                'meta': '../data/COVID19_ICU',
                'adj': '../data/COVID19_ICU'}

        return dirs


class Covid19ICULength(DatasetSetupBuilder):
    def __init__(self):
        super(Covid19ICULength, self).__init__()

    def build_dataset(self):
        general_dir = '../data/COVID19_ICU_Length'
        if not os.path.isdir(general_dir):
            print('Creating COVID19_ICU_Length folder structure ...')
            file = '../sources/COVID19/COVID19_Death.xlsx'
            data = load_workbook(file)
            data_work = data['COVID19']

            feature_list = ['B', 'R', 'Y', 'AH', 'AI', 'AN', 'AQ', 'AT', 'AW', 'AZ', 'BC',
                            'BF', 'BI']

            meta_list = ['B', 'AH', 'AI', 'AJ']

            target = 'L'

            features = []
            meta_data = []
            targets = []
            feature_names = []

            for j, para in enumerate(feature_list):
                feature_names.append(data_work[para + str(1)].value)

            idx = 2
            while data_work['A' + str(idx)].value is not None:
                curr_feat = []
                curr_meta = []
                if data_work[target + str(idx)].value is not None:
                    targets.append(data_work[target + str(idx)].value)
                else:
                    idx += 1
                    continue
                for j, para in enumerate(feature_list):
                    if data_work[para + str(idx)].value is not None:
                        val = data_work[para + str(idx)].value
                        curr_feat.append(np.round(val, 3))
                    else:
                        curr_feat.append(-1)
                features.append(curr_feat)
                for j, para in enumerate(meta_list):
                    if data_work[para + str(idx)].value is not None and not data_work[para + str(idx)].value == -1:
                        curr_meta.append(data_work[para + str(idx)].value)
                    else:
                        curr_meta.append(-1)
                meta_data.append(curr_meta)
                idx += 1

            features = np.array(features)
            meta_data = np.array(meta_data)

            for i in range(features.shape[1]):
                mask = features[:, i] == -1
                features[mask, i] = features.mean(0)[i]

            for i in range(meta_data.shape[1]):
                mask = meta_data[:, i] == -1
                meta_data[mask, i] = meta_data.mean(0)[i]

            create_dir('../data')
            create_dir(general_dir)
            np.save(general_dir + '/features.npy', features)
            np.save(general_dir + '/targets.npy', targets)
            np.save(general_dir + '/meta_data.npy', meta_data)
            np.save(general_dir + '/feature_names.npy', feature_names)

            print('COVID19_Death folder structure created, proceeding ...')
        else:
            print('COVID19_Death folder structure already created, proceeding ...')

        dirs = {'img': None,
                'feat': general_dir,
                'targets': [general_dir],
                'loss': [torch.nn.MSELoss],
                'meta': general_dir,
                'adj': general_dir}
        return dirs


class Covid19DeathLength(DatasetSetupBuilder):
    def __init__(self):
        super(Covid19DeathLength, self).__init__()

    def build_dataset(self):
        general_dir = '../data/COVID19_Death_Length'
        if not os.path.isdir(general_dir):
            print('Creating COVID19_ICU_Length folder structure ...')
            file = '../sources/COVID19/COVID19_Death.xlsx'
            data = load_workbook(file)
            data_work = data['COVID19']

            feature_list = ['B', 'R', 'Y', 'AH', 'AI', 'AN', 'AQ', 'AT', 'AW', 'AZ', 'BC',
                            'BF', 'BI']

            meta_list = ['B', 'AH', 'AI', 'AJ']

            target = 'CA'

            features = []
            meta_data = []
            targets = []

            idx = 2
            while data_work['A' + str(idx)].value is not None:
                curr_feat = []
                curr_meta = []
                if data_work[target + str(idx)].value is not None:
                    targets.append(data_work[target + str(idx)].value)
                else:
                    idx += 1
                    continue
                for j, para in enumerate(feature_list):
                    if data_work[para + str(idx)].value is not None:
                        val = data_work[para + str(idx)].value
                        curr_feat.append(np.round(val, 3))
                    else:
                        curr_feat.append(-1)
                features.append(curr_feat)
                for j, para in enumerate(meta_list):
                    if data_work[para + str(idx)].value is not None and not data_work[para + str(idx)].value == -1:
                        curr_meta.append(data_work[para + str(idx)].value)
                    else:
                        curr_meta.append(-1)
                meta_data.append(curr_meta)
                idx += 1

            features = np.array(features)
            meta_data = np.array(meta_data)

            for i in range(features.shape[1]):
                mask = features[:, i] == -1
                features[mask, i] = features.mean(0)[i]

            for i in range(meta_data.shape[1]):
                mask = meta_data[:, i] == -1
                meta_data[mask, i] = meta_data.mean(0)[i]

            create_dir('../data')
            create_dir(general_dir)
            np.save(general_dir + '/features.npy', features)
            np.save(general_dir + '/targets.npy', targets)
            np.save(general_dir + '/meta_data.npy', meta_data)
            print('COVID19_Death folder structure created, proceeding ...')
        else:
            print('COVID19_Death folder structure already created, proceeding ...')

        dirs = {'img': None,
                'feat': general_dir,
                'targets': [general_dir],
                'loss': [torch.nn.MSELoss],
                'meta': general_dir,
                'adj': general_dir}
        return dirs


class Covid19Death(DatasetSetupBuilder):
    def __init__(self):
        super(Covid19Death, self).__init__()

    def build_dataset(self):
        if not os.path.isdir('../data/COVID19_Death'):
            print('Creating COVID19_Death folder structure ...')
            file = '../sources/COVID19/COVID19_Death.xlsx'
            data = load_workbook(file)
            data_work = data['COVID19']

            feature_list = ['B', 'R', 'Y', 'AH', 'AI', 'AN', 'AQ', 'AT', 'AW', 'AZ', 'BC',
                            'BF', 'BI']

            meta_list = ['B', 'AH', 'AI', 'AJ']

            target = 'BY'

            features = []
            meta_data = []
            targets = []
            feature_names = []

            for j, para in enumerate(feature_list):
                feature_names.append(data_work[para + str(1)].value)

            idx = 2
            while data_work['A' + str(idx)].value is not None:
                curr_feat = []
                curr_meta = []
                if data_work[target + str(idx)].value is not None:
                    targets.append(data_work[target + str(idx)].value)
                else:
                    idx += 1
                    continue
                for j, para in enumerate(feature_list):
                    if data_work[para + str(idx)].value is not None:
                        val = data_work[para + str(idx)].value
                        curr_feat.append(np.round(val, 3))
                    else:
                        curr_feat.append(-1)
                features.append(curr_feat)
                for j, para in enumerate(meta_list):
                    if data_work[para + str(idx)].value is not None and not data_work[para + str(idx)].value == -1:
                        curr_meta.append(data_work[para + str(idx)].value)
                    else:
                        curr_meta.append(-1)
                meta_data.append(curr_meta)
                idx += 1

            features = np.array(features)
            meta_data = np.array(meta_data)

            for i in range(features.shape[1]):
                mask = features[:, i] == -1
                features[mask, i] = features.mean(0)[i]

            for i in range(meta_data.shape[1]):
                mask = meta_data[:, i] == -1
                meta_data[mask, i] = meta_data.mean(0)[i]

            create_dir('../data')
            create_dir('../data/COVID19_Death')
            np.save('../data/COVID19_Death/features.npy', features)
            np.save('../data/COVID19_Death/targets.npy', targets)
            np.save('../data/COVID19_Death/meta_data.npy', meta_data)
            np.save('../data/COVID19_Death/feature_names.npy', feature_names)
            print('COVID19_Death folder structure created, proceeding ...')
        else:
            print('COVID19_Death folder structure already created, proceeding ...')

        dirs = {'img': None,
                'feat': '../data/COVID19_Death',
                'targets': ['../data/COVID19_Death'],
                'loss': [torch.nn.BCEWithLogitsLoss],
                'meta': '../data/COVID19_Death',
                'adj': '../data/COVID19_Death'}
        return dirs
